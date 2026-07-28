from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def generate_excel_report(queryset):
    """
    Generate an Excel workbook binary stream for attendance records.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Header styling
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    headers = [
        "S.No", "Date", "Roll Number", "Student Name", 
        "Department", "Entry Time", "Exit Time", "Status", "Confidence (%)", "Entry Location"
    ]
    
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    row_font = Font(name="Segoe UI", size=10)
    
    for idx, att in enumerate(queryset, start=1):
        entry_str = att.entry_time.strftime("%I:%M %p") if att.entry_time else "--"
        exit_str = att.exit_time.strftime("%I:%M %p") if att.exit_time else "--"
        confidence_str = f"{att.confidence:.1f}%" if att.confidence else "N/A"
        loc_str = f"{att.entry_lat:.4f}, {att.entry_lon:.4f}" if att.entry_lat else "N/A"

        row_data = [
            idx,
            att.date.strftime("%Y-%m-%d"),
            att.student.roll_number,
            att.student.user.get_full_name(),
            att.student.department.name if att.student.department else "N/A",
            entry_str,
            exit_str,
            att.get_status_display(),
            confidence_str,
            loc_str
        ]
        ws.append(row_data)

        # Style data row
        row_num = idx + 1
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = row_font
            cell.border = thin_border
            if col_num in [1, 2, 6, 7, 8, 9]:
                cell.alignment = align_center

    # Adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()


def generate_pdf_report(queryset, title="Smart Attendance System Report"):
    """
    Generate a PDF document binary stream for attendance records using ReportLab.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=40, bottomMargin=30)
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#4F46E5'),
        spaceAfter=15,
        alignment=1 # Center
    )

    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 10))

    headers = ['Date', 'Roll No', 'Name', 'Department', 'Entry', 'Exit', 'Status']
    table_data = [headers]

    cell_style = ParagraphStyle('Cell', fontName='Helvetica', fontSize=8, leading=10)
    header_style = ParagraphStyle('Header', fontName='Helvetica-Bold', fontSize=9, textColor=colors.white, leading=11)

    for att in queryset:
        entry_str = att.entry_time.strftime("%I:%M %p") if att.entry_time else "--"
        exit_str = att.exit_time.strftime("%I:%M %p") if att.exit_time else "--"
        
        table_data.append([
            Paragraph(att.date.strftime("%Y-%m-%d"), cell_style),
            Paragraph(att.student.roll_number, cell_style),
            Paragraph(att.student.user.get_full_name(), cell_style),
            Paragraph(att.student.department.name if att.student.department else "-", cell_style),
            Paragraph(entry_str, cell_style),
            Paragraph(exit_str, cell_style),
            Paragraph(att.get_status_display(), cell_style)
        ])

    # Build Table
    t = Table(table_data, colWidths=[65, 75, 120, 110, 55, 55, 60])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
    ]))

    elements.append(t)
    doc.build(elements)

    buffer.seek(0)
    return buffer.getvalue()
